import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from uuid import uuid4
from zipfile import BadZipFile, ZipFile, is_zipfile

import duckdb
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from backend.app.core.config import Settings
from backend.app.core.errors import InvalidSourceError
from backend.app.domain.models import (
    ColumnProfile,
    ProfileOptions,
    SheetProfile,
    SourcePreview,
)
from backend.app.profiling.header_detection import (
    build_column_names,
    detect_header,
    infer_data_type,
    is_null,
    normalize_text,
)


def _trim_trailing_nulls(values: list[Any]) -> list[Any]:
    normalized = list(values)
    while normalized and is_null(normalized[-1]):
        normalized.pop()
    return normalized


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if is_null(value):
        return None
    return value


class XlsxProfiler:
    """Streaming XLSX profiler with disk-backed duplicate counting."""

    name = "openpyxl+duckdb"
    supported_extensions = frozenset({".xlsx"})

    def __init__(self, settings: Settings) -> None:
        self._settings = settings

    def profile(
        self,
        path: Path,
        *,
        source_id: str,
        options: ProfileOptions,
    ) -> list[SheetProfile]:
        self._validate_container(path)

        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            ValueError,
        ) as exc:
            raise InvalidSourceError(
                "The XLSX workbook could not be opened.",
                code="xlsx_open_failed",
            ) from exc

        try:
            return [
                self._profile_sheet(
                    worksheet,
                    source_id=source_id,
                    explicit_header=options.header_for(worksheet.title),
                )
                for worksheet in workbook.worksheets
            ]
        finally:
            workbook.close()

    def preview(
        self,
        path: Path,
        *,
        source_id: str,
        sheet: SheetProfile,
        offset: int,
        limit: int,
    ) -> SourcePreview:
        self._validate_container(path)
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            ValueError,
        ) as exc:
            raise InvalidSourceError(
                "The XLSX workbook could not be opened for preview.",
                code="xlsx_preview_open_failed",
            ) from exc

        try:
            if sheet.name not in workbook.sheetnames:
                raise InvalidSourceError(
                    "The requested sheet does not exist.",
                    code="source_sheet_not_found",
                    status_code=404,
                    details={"sheet": sheet.name},
                )
            worksheet = workbook[sheet.name]
            columns = [column.name for column in sheet.columns]
            header_row = sheet.header_row_number
            if header_row is None or not columns:
                return SourcePreview(
                    source_id=source_id,
                    sheet_name=sheet.name,
                    columns=columns,
                    rows=[],
                    offset=offset,
                    limit=limit,
                    returned=0,
                    total_rows=sheet.row_count,
                )

            preview_rows: list[list[str | None]] = []
            data_index = 0
            for physical_row, raw_row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                if physical_row <= header_row:
                    continue
                values = _trim_trailing_nulls(list(raw_row))
                if not values:
                    continue
                if data_index < offset:
                    data_index += 1
                    continue

                padded = values[: len(columns)] + [None] * max(
                    0, len(columns) - len(values)
                )
                preview_rows.append(
                    [
                        None if is_null(value) else normalize_text(value)
                        for value in padded
                    ]
                )
                data_index += 1
                if len(preview_rows) >= limit:
                    break

            return SourcePreview(
                source_id=source_id,
                sheet_name=sheet.name,
                columns=columns,
                rows=preview_rows,
                offset=offset,
                limit=limit,
                returned=len(preview_rows),
                total_rows=sheet.row_count,
            )
        finally:
            workbook.close()

    def _validate_container(self, path: Path) -> None:
        if not is_zipfile(path):
            raise InvalidSourceError(
                "The XLSX file is not a valid Open XML workbook.",
                code="invalid_xlsx_container",
            )

        try:
            with ZipFile(path) as archive:
                total_uncompressed = 0
                for member in archive.infolist():
                    if member.flag_bits & 0x1:
                        raise InvalidSourceError(
                            "Encrypted XLSX workbooks are not supported.",
                            code="encrypted_xlsx_not_supported",
                        )

                    total_uncompressed += member.file_size
                    if total_uncompressed > self._settings.xlsx_max_uncompressed_bytes:
                        raise InvalidSourceError(
                            "The XLSX workbook expands beyond the configured limit.",
                            code="xlsx_uncompressed_size_exceeded",
                            status_code=413,
                            details={
                                "max_uncompressed_bytes": (
                                    self._settings.xlsx_max_uncompressed_bytes
                                )
                            },
                        )

                    compressed_size = max(member.compress_size, 1)
                    ratio = member.file_size / compressed_size
                    if ratio > self._settings.xlsx_max_compression_ratio:
                        raise InvalidSourceError(
                            "The XLSX workbook has a suspicious compression ratio.",
                            code="xlsx_suspicious_compression",
                            status_code=422,
                            details={
                                "member": member.filename,
                                "compression_ratio": round(ratio, 2),
                                "max_compression_ratio": (
                                    self._settings.xlsx_max_compression_ratio
                                ),
                            },
                        )
        except BadZipFile as exc:
            raise InvalidSourceError(
                "The XLSX file is not a valid Open XML workbook.",
                code="invalid_xlsx_container",
            ) from exc

    def _profile_sheet(
        self,
        worksheet: Any,
        *,
        source_id: str,
        explicit_header: int | None,
    ) -> SheetProfile:
        sampled_rows: list[list[Any]] = []
        for row in worksheet.iter_rows(values_only=True):
            sampled_rows.append(_trim_trailing_nulls(list(row)))
            if len(sampled_rows) >= self._settings.header_scan_rows:
                break

        detection = detect_header(sampled_rows)
        header_row = explicit_header or detection.row_number

        if header_row is None:
            return SheetProfile(
                name=worksheet.title,
                header_row_number=None,
                header_confidence=None,
                header_candidates=[],
                row_count=0,
                column_count=0,
                duplicate_row_count=0,
                total_cell_count=0,
                null_cell_count=0,
                null_percentage=0.0,
                columns=[],
            )

        rows = worksheet.iter_rows(values_only=True)
        header_values: list[Any] = []
        column_names: list[str] = []
        null_counts: list[int] = []
        non_null_counts: list[int] = []
        samples: list[list[Any]] = []
        row_count = 0

        work_dir = self._settings.work_storage / source_id
        work_dir.mkdir(parents=True, exist_ok=True)
        duplicate_db = work_dir / f"duplicates-{uuid4().hex}.duckdb"
        connection = duckdb.connect(str(duplicate_db))
        escaped = work_dir.as_posix().replace("'", "''")
        connection.execute(f"SET temp_directory='{escaped}'")
        connection.execute(f"SET memory_limit='{self._settings.duckdb_memory_limit}'")
        connection.execute(f"SET threads={self._settings.duckdb_threads}")
        connection.execute("CREATE TABLE row_keys (row_key VARCHAR)")
        batch: list[tuple[str]] = []

        try:
            for physical_row, raw_row in enumerate(rows, start=1):
                values = _trim_trailing_nulls(list(raw_row))

                if physical_row < header_row:
                    continue
                if physical_row == header_row:
                    header_values = values
                    column_names = build_column_names(
                        header_values,
                        len(header_values),
                    )
                    null_counts = [0] * len(column_names)
                    non_null_counts = [0] * len(column_names)
                    samples = [[] for _ in column_names]
                    continue
                if not values:
                    continue

                if len(values) > len(column_names):
                    additional = len(values) - len(column_names)
                    column_names = build_column_names(
                        header_values,
                        len(values),
                    )
                    null_counts.extend([row_count] * additional)
                    non_null_counts.extend([0] * additional)
                    samples.extend([[] for _ in range(additional)])

                padded = values + [None] * (len(column_names) - len(values))
                row_count += 1

                for index, value in enumerate(padded):
                    if is_null(value):
                        null_counts[index] += 1
                    else:
                        non_null_counts[index] += 1
                        if len(samples[index]) < self._settings.profile_sample_values:
                            normalized = normalize_text(value)
                            existing = {normalize_text(item) for item in samples[index]}
                            if normalized not in existing:
                                samples[index].append(value)

                row_key = json.dumps(
                    [_json_safe(value) for value in padded],
                    ensure_ascii=False,
                    separators=(",", ":"),
                    default=str,
                )
                batch.append((row_key,))
                if len(batch) >= 5000:
                    connection.executemany(
                        "INSERT INTO row_keys VALUES (?)",
                        batch,
                    )
                    batch.clear()

            if batch:
                connection.executemany(
                    "INSERT INTO row_keys VALUES (?)",
                    batch,
                )

            duplicate_row_count = int(
                connection.execute(
                    "SELECT COUNT(*) - COUNT(DISTINCT row_key) FROM row_keys"
                ).fetchone()[0]
                or 0
            )
        finally:
            connection.close()
            duplicate_db.unlink(missing_ok=True)
            duplicate_db.with_suffix(".duckdb.wal").unlink(missing_ok=True)

        if not column_names:
            raise InvalidSourceError(
                "The selected header row does not exist in the sheet.",
                code="header_row_out_of_range",
                details={
                    "sheet": worksheet.title,
                    "header_row_number": header_row,
                },
            )

        columns = [
            ColumnProfile(
                name=column_names[index],
                position=index + 1,
                data_type=infer_data_type(samples[index]),
                null_count=null_counts[index],
                non_null_count=non_null_counts[index],
                null_percentage=(
                    round(
                        (null_counts[index] / row_count) * 100,
                        4,
                    )
                    if row_count
                    else 0.0
                ),
                distinct_count=None,
                sample_values=[normalize_text(value) for value in samples[index]],
            )
            for index in range(len(column_names))
        ]

        total_nulls = sum(null_counts)
        total_cells = row_count * len(column_names)

        return SheetProfile(
            name=worksheet.title,
            header_row_number=header_row,
            header_confidence=(1.0 if explicit_header else detection.confidence),
            header_candidates=detection.candidates,
            row_count=row_count,
            column_count=len(column_names),
            duplicate_row_count=duplicate_row_count,
            total_cell_count=total_cells,
            null_cell_count=total_nulls,
            null_percentage=(
                round((total_nulls / total_cells) * 100, 4) if total_cells else 0.0
            ),
            columns=columns,
        )
