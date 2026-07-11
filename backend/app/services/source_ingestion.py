import csv
import hashlib
import json
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4
from zipfile import BadZipFile, is_zipfile

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from backend.app.schemas.source import (
    ColumnProfile,
    SheetProfile,
    SourceProfile,
)

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
MAX_FILE_SIZE_BYTES = 25 * 1024 * 1024
UPLOAD_CHUNK_SIZE = 1024 * 1024

BACKEND_ROOT = Path(__file__).resolve().parents[2]
STORAGE_ROOT = BACKEND_ROOT / "storage"
SOURCE_STORAGE = STORAGE_ROOT / "sources"
PROFILE_STORAGE = STORAGE_ROOT / "profiles"


class SourceIngestionError(Exception):
    def __init__(self, message: str, status_code: int = 400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code


def _is_null(value: Any) -> bool:
    return value is None or (
        isinstance(value, str) and not value.strip()
    )


def _trim_trailing_nulls(values: Iterable[Any]) -> list[Any]:
    normalized = list(values)

    while normalized and _is_null(normalized[-1]):
        normalized.pop()

    return normalized


def _normalize_for_duplicate_key(value: Any) -> str:
    if _is_null(value):
        return "<NULL>"

    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    return str(value).strip()


def _build_column_names(
    header_values: Iterable[Any],
    width: int,
) -> list[str]:
    raw_values = list(header_values)
    names: list[str] = []
    occurrences: dict[str, int] = {}

    for index in range(width):
        value = raw_values[index] if index < len(raw_values) else None
        base_name = (
            str(value).strip()
            if not _is_null(value)
            else f"column_{index + 1}"
        )

        occurrence = occurrences.get(base_name, 0) + 1
        occurrences[base_name] = occurrence

        names.append(
            base_name
            if occurrence == 1
            else f"{base_name}_{occurrence}"
        )

    return names


def _profile_rows(
    sheet_name: str,
    rows: Iterable[Iterable[Any]],
) -> SheetProfile:
    header_values: list[Any] | None = None
    header_row_number: int | None = None
    column_names: list[str] = []
    null_counts: list[int] = []
    non_null_counts: list[int] = []
    row_count = 0
    duplicate_row_count = 0
    seen_rows: set[tuple[str, ...]] = set()

    for physical_row_number, row in enumerate(rows, start=1):
        values = _trim_trailing_nulls(row)

        if not values:
            continue

        if header_values is None:
            header_values = values
            header_row_number = physical_row_number
            column_names = _build_column_names(
                header_values,
                len(header_values),
            )
            null_counts = [0] * len(header_values)
            non_null_counts = [0] * len(header_values)
            continue

        if len(values) > len(column_names):
            additional_columns = len(values) - len(column_names)
            column_names = _build_column_names(
                header_values,
                len(values),
            )
            null_counts.extend([row_count] * additional_columns)
            non_null_counts.extend([0] * additional_columns)

        padded_row = values + [None] * (
            len(column_names) - len(values)
        )
        row_count += 1

        for index, value in enumerate(padded_row):
            if _is_null(value):
                null_counts[index] += 1
            else:
                non_null_counts[index] += 1

        duplicate_key = tuple(
            _normalize_for_duplicate_key(value)
            for value in _trim_trailing_nulls(padded_row)
        )

        if duplicate_key in seen_rows:
            duplicate_row_count += 1
        else:
            seen_rows.add(duplicate_key)

    if header_values is None:
        return SheetProfile(
            name=sheet_name,
            header_row_number=None,
            row_count=0,
            column_count=0,
            duplicate_row_count=0,
            columns=[],
        )

    columns = [
        ColumnProfile(
            name=column_names[index],
            position=index + 1,
            null_count=null_counts[index],
            non_null_count=non_null_counts[index],
        )
        for index in range(len(column_names))
    ]

    return SheetProfile(
        name=sheet_name,
        header_row_number=header_row_number,
        row_count=row_count,
        column_count=len(column_names),
        duplicate_row_count=duplicate_row_count,
        columns=columns,
    )


def _profile_csv_with_encoding(
    path: Path,
    encoding: str,
) -> SheetProfile:
    with path.open(
        "r",
        encoding=encoding,
        errors="strict",
        newline="",
    ) as source_file:
        sample = source_file.read(65536)

        if "\x00" in sample:
            raise SourceIngestionError(
                "The CSV file contains binary null bytes.",
                status_code=422,
            )

        source_file.seek(0)

        try:
            dialect = csv.Sniffer().sniff(
                sample,
                delimiters=",;\t|",
            )
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(source_file, dialect)
        return _profile_rows(path.stem, reader)


def _profile_csv(path: Path) -> list[SheetProfile]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return [
                _profile_csv_with_encoding(
                    path,
                    encoding,
                )
            ]
        except UnicodeDecodeError:
            continue

    raise SourceIngestionError(
        "The CSV encoding is not supported. "
        "Use UTF-8 or Windows-1252.",
        status_code=422,
    )


def _profile_xlsx(path: Path) -> list[SheetProfile]:
    if not is_zipfile(path):
        raise SourceIngestionError(
            "The XLSX file is not a valid Open XML workbook.",
            status_code=422,
        )

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
        )
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as exc:
        raise SourceIngestionError(
            "The XLSX workbook could not be opened.",
            status_code=422,
        ) from exc

    try:
        return [
            _profile_rows(
                worksheet.title,
                worksheet.iter_rows(values_only=True),
            )
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def profile_source_path(
    path: Path,
    extension: str,
) -> list[SheetProfile]:
    normalized_extension = extension.lower()

    if normalized_extension == ".csv":
        return _profile_csv(path)

    if normalized_extension == ".xlsx":
        return _profile_xlsx(path)

    raise SourceIngestionError(
        "Only CSV and XLSX files are supported.",
        status_code=400,
    )


def _safe_original_name(filename: str | None) -> str:
    if not filename:
        raise SourceIngestionError(
            "The uploaded file must have a name.",
            status_code=400,
        )

    normalized = filename.replace("\\", "/")
    safe_name = Path(normalized).name.strip()

    if not safe_name:
        raise SourceIngestionError(
            "The uploaded file name is invalid.",
            status_code=400,
        )

    return safe_name


async def _store_upload(
    upload: UploadFile,
    destination: Path,
) -> tuple[int, str]:
    temporary_destination = destination.with_suffix(
        destination.suffix + ".part"
    )
    size_bytes = 0
    digest = hashlib.sha256()

    try:
        with temporary_destination.open("xb") as stored_file:
            while chunk := await upload.read(UPLOAD_CHUNK_SIZE):
                size_bytes += len(chunk)

                if size_bytes > MAX_FILE_SIZE_BYTES:
                    raise SourceIngestionError(
                        "The file exceeds the 25 MB limit.",
                        status_code=413,
                    )

                digest.update(chunk)
                stored_file.write(chunk)

        if size_bytes == 0:
            raise SourceIngestionError(
                "The uploaded file is empty.",
                status_code=400,
            )

        temporary_destination.replace(destination)
        return size_bytes, digest.hexdigest()
    except Exception:
        temporary_destination.unlink(missing_ok=True)
        destination.unlink(missing_ok=True)
        raise
    finally:
        await upload.close()


def _persist_profile(source_profile: SourceProfile) -> None:
    PROFILE_STORAGE.mkdir(parents=True, exist_ok=True)
    destination = PROFILE_STORAGE / f"{source_profile.id}.json"
    temporary_destination = destination.with_suffix(".json.tmp")

    payload = json.dumps(
        source_profile.model_dump(mode="json"),
        ensure_ascii=False,
        indent=2,
    )

    temporary_destination.write_text(
        payload,
        encoding="utf-8",
    )
    temporary_destination.replace(destination)


async def ingest_source_file(
    upload: UploadFile,
) -> SourceProfile:
    original_name = _safe_original_name(upload.filename)
    extension = Path(original_name).suffix.lower()

    if extension not in ALLOWED_EXTENSIONS:
        raise SourceIngestionError(
            "Only CSV and XLSX files are supported.",
            status_code=400,
        )

    source_id = uuid4().hex
    SOURCE_STORAGE.mkdir(parents=True, exist_ok=True)
    stored_path = SOURCE_STORAGE / f"{source_id}{extension}"

    try:
        size_bytes, sha256 = await _store_upload(
            upload,
            stored_path,
        )
        sheets = profile_source_path(
            stored_path,
            extension,
        )

        source_profile = SourceProfile(
            id=source_id,
            original_name=original_name,
            extension=extension.lstrip("."),
            media_type=upload.content_type,
            size_bytes=size_bytes,
            sha256=sha256,
            stored_at=datetime.now(timezone.utc),
            sheets=sheets,
        )

        _persist_profile(source_profile)
        return source_profile
    except Exception:
        stored_path.unlink(missing_ok=True)
        raise
