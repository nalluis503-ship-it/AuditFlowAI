from __future__ import annotations

import shutil
from pathlib import Path
from uuid import uuid4

import duckdb
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from backend.app.core.config import Settings
from backend.app.core.errors import InvalidSourceError
from backend.app.profiling.csv_format import inspect_csv_format
from backend.app.profiling.header_detection import is_null, normalize_text
from backend.app.tabular.compiler import (
    TabularPlanCompiler,
    quote_identifier,
    quote_literal,
)
from backend.app.tabular.contracts import (
    CancellationCheck,
    MaterializedTabularResult,
    ProgressReporter,
    ResolvedTabularInput,
)


class DuckDBTabularEngine:
    """Out-of-core typed-plan execution that materializes derived Parquet sources."""

    name = "duckdb"

    def __init__(self, settings: Settings) -> None:
        self._settings = settings
        self._compiler = TabularPlanCompiler()

    def materialize(
        self,
        plan,
        *,
        run_id: str,
        inputs: list[ResolvedTabularInput],
        report_progress: ProgressReporter,
        check_cancelled: CancellationCheck,
    ) -> MaterializedTabularResult:
        work_dir = self._work_dir(run_id)
        work_dir.mkdir(parents=True, exist_ok=True)
        output_path = work_dir / "result.parquet"
        output_path.unlink(missing_ok=True)

        connection = self._connect(work_dir)
        try:
            input_views: dict[str, str] = {}
            input_schemas: dict[str, list[str]] = {}
            total = max(len(inputs), 1)

            for index, item in enumerate(inputs, start=1):
                check_cancelled()
                report_progress(
                    10.0 + (20.0 * ((index - 1) / total)),
                    "preparing_inputs",
                    f"Preparing tabular input {index} of {len(inputs)}.",
                )
                relation = self._open_relation(
                    connection,
                    item,
                    work_dir=work_dir,
                    input_position=index,
                )
                view_name = f"input_{index}_{uuid4().hex[:8]}"
                relation.create_view(view_name, replace=True)
                columns = list(relation.columns)
                expected = [column.name for column in item.sheet.columns]
                if columns != expected:
                    raise InvalidSourceError(
                        "The physical source schema no longer matches its profile.",
                        code="tabular_source_schema_changed",
                        status_code=409,
                        details={
                            "source_id": item.source.id,
                            "sheet": item.sheet.name,
                            "profile_columns": expected,
                            "physical_columns": columns,
                        },
                    )
                input_views[item.alias] = view_name
                input_schemas[item.alias] = columns

            check_cancelled()
            report_progress(
                35.0,
                "compiling_plan",
                "Validating and compiling the typed tabular plan.",
            )
            compiled = self._compiler.compile(
                plan,
                input_views=input_views,
                input_schemas=input_schemas,
            )

            check_cancelled()
            report_progress(
                45.0,
                "executing_plan",
                "Executing the tabular plan with bounded memory.",
            )
            output_literal = quote_literal(output_path.as_posix())
            try:
                parquet_query = self._parquet_safe_query(
                    connection,
                    compiled.sql,
                    compiled.output_columns,
                )
                connection.execute(
                    f"COPY ({parquet_query}) TO {output_literal} "
                    "(FORMAT PARQUET, COMPRESSION ZSTD)"
                )
            except (
                duckdb.BinderException,
                duckdb.CatalogException,
                duckdb.ConversionException,
                duckdb.InvalidInputException,
                duckdb.ParserException,
            ) as exc:
                raise InvalidSourceError(
                    "The typed tabular plan is incompatible with the source data.",
                    code="tabular_plan_execution_invalid",
                    status_code=422,
                    details={"engine": self.name},
                ) from exc
            except (duckdb.IOException, duckdb.OutOfMemoryException) as exc:
                raise InvalidSourceError(
                    "The tabular engine could not materialize the result with "
                    "the available runtime resources.",
                    code="tabular_engine_resource_failure",
                    status_code=503,
                    details={"engine": self.name},
                ) from exc

            check_cancelled()
            if not output_path.is_file():
                raise InvalidSourceError(
                    "The tabular engine did not produce an output artifact.",
                    code="tabular_output_missing",
                    status_code=500,
                )
            try:
                metadata = pq.read_metadata(output_path)
            except Exception as exc:
                raise InvalidSourceError(
                    "The derived Parquet output could not be verified.",
                    code="tabular_output_invalid",
                    status_code=500,
                ) from exc
            physical_columns = metadata.schema.names
            if physical_columns != list(compiled.output_columns):
                raise InvalidSourceError(
                    "The derived output schema differs from the compiled plan.",
                    code="tabular_output_schema_mismatch",
                    status_code=500,
                    details={
                        "compiled_columns": list(compiled.output_columns),
                        "physical_columns": physical_columns,
                    },
                )

            report_progress(
                70.0,
                "output_ready",
                "The derived Parquet artifact is ready for registration.",
            )
            return MaterializedTabularResult(
                path=output_path,
                engine=self.name,
                output_columns=compiled.output_columns,
            )
        finally:
            connection.close()

    @staticmethod
    def _parquet_safe_query(
        connection: duckdb.DuckDBPyConnection,
        sql: str,
        output_columns: tuple[str, ...],
    ) -> str:
        """Preserve exact HUGEINT results when DuckDB writes Parquet.

        DuckDB promotes SUM(BIGINT) to HUGEINT. Parquet has no native signed
        128-bit integer, and DuckDB otherwise serializes that value as DOUBLE,
        which can silently lose precision. DECIMAL(38, 0) is exact and
        supported by Parquet. Values outside its range fail explicitly instead
        of being written as an imprecise floating-point number.
        """
        relation = connection.sql(sql)
        physical_columns = list(relation.columns)
        if physical_columns != list(output_columns):
            raise InvalidSourceError(
                "The tabular result schema differs from the compiled plan.",
                code="tabular_result_schema_mismatch",
                status_code=500,
                details={
                    "compiled_columns": list(output_columns),
                    "result_columns": physical_columns,
                },
            )

        result_types = [str(data_type).upper() for data_type in relation.types]
        if "HUGEINT" not in result_types:
            return sql

        view_name = f"tabular_result_{uuid4().hex[:12]}"
        relation.create_view(view_name, replace=True)
        projections: list[str] = []
        for column, data_type in zip(
            output_columns,
            result_types,
            strict=True,
        ):
            quoted = quote_identifier(column)
            if data_type == "HUGEINT":
                projections.append(f"CAST({quoted} AS DECIMAL(38, 0)) AS {quoted}")
            else:
                projections.append(quoted)
        return f"SELECT {', '.join(projections)} FROM {quote_identifier(view_name)}"

    def cleanup(self, run_id: str) -> None:
        work_dir = self._work_dir(run_id)
        if work_dir.exists():
            shutil.rmtree(work_dir, ignore_errors=True)

    def _work_dir(self, run_id: str) -> Path:
        return self._settings.work_storage / "tabular" / run_id

    def _connect(self, work_dir: Path) -> duckdb.DuckDBPyConnection:
        connection = duckdb.connect(database=":memory:")
        escaped = work_dir.as_posix().replace("'", "''")
        connection.execute(f"SET temp_directory='{escaped}'")
        connection.execute(f"SET memory_limit='{self._settings.duckdb_memory_limit}'")
        connection.execute(f"SET threads={self._settings.duckdb_threads}")
        return connection

    def _open_relation(
        self,
        connection: duckdb.DuckDBPyConnection,
        item: ResolvedTabularInput,
        *,
        work_dir: Path,
        input_position: int,
    ) -> duckdb.DuckDBPyRelation:
        if not item.source.stored_path:
            raise InvalidSourceError(
                "A tabular input does not have a stored physical file.",
                code="source_file_missing",
                details={"source_id": item.source.id},
            )
        path = Path(item.source.stored_path)
        if not path.is_file():
            raise InvalidSourceError(
                "A tabular input physical file is missing.",
                code="source_file_missing",
                details={"source_id": item.source.id},
            )

        extension = path.suffix.lower()
        try:
            if extension == ".csv":
                csv_format = inspect_csv_format(
                    path,
                    scan_rows=self._settings.header_scan_rows,
                )
                return connection.read_csv(
                    str(path),
                    header=True,
                    delimiter=csv_format.delimiter,
                    skiprows=(item.sheet.header_row_number or 1) - 1,
                    all_varchar=True,
                    encoding=csv_format.duckdb_encoding,
                )
            if extension == ".parquet":
                return connection.read_parquet(str(path))
            if extension == ".xlsx":
                staged = work_dir / f"xlsx-{input_position}.parquet"
                self._stage_xlsx(item, path=path, destination=staged)
                return connection.read_parquet(str(staged))
        except InvalidSourceError:
            raise
        except Exception as exc:
            raise InvalidSourceError(
                "The tabular input could not be opened by the execution engine.",
                code="tabular_input_open_failed",
                details={
                    "source_id": item.source.id,
                    "extension": extension,
                    "sheet": item.sheet.name,
                },
            ) from exc

        raise InvalidSourceError(
            "The tabular execution engine does not support this source format.",
            code="unsupported_tabular_input_format",
            details={"extension": extension},
        )

    def _stage_xlsx(
        self,
        item: ResolvedTabularInput,
        *,
        path: Path,
        destination: Path,
    ) -> None:
        destination.unlink(missing_ok=True)
        columns = [column.name for column in item.sheet.columns]
        schema = pa.schema([pa.field(column, pa.string()) for column in columns])
        writer = pq.ParquetWriter(destination, schema=schema, compression="zstd")
        workbook = None
        wrote_rows = False
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
            if item.sheet.name not in workbook.sheetnames:
                raise InvalidSourceError(
                    "The requested XLSX sheet no longer exists.",
                    code="source_sheet_not_found",
                    status_code=409,
                    details={"sheet": item.sheet.name},
                )
            worksheet = workbook[item.sheet.name]
            header_row = item.sheet.header_row_number
            if header_row is None:
                raise InvalidSourceError(
                    "The XLSX sheet does not have a confirmed header row.",
                    code="source_header_required",
                    status_code=409,
                    details={"sheet": item.sheet.name},
                )

            batch: dict[str, list[str | None]] = {column: [] for column in columns}
            batch_rows = 0
            for physical_row, raw_row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                if physical_row <= header_row:
                    continue
                values = list(raw_row[: len(columns)])
                if not values or all(is_null(value) for value in values):
                    continue
                values += [None] * (len(columns) - len(values))
                for index, column in enumerate(columns):
                    value = values[index]
                    batch[column].append(
                        None if is_null(value) else normalize_text(value)
                    )
                batch_rows += 1
                if batch_rows >= 5000:
                    writer.write_table(pa.table(batch, schema=schema))
                    wrote_rows = True
                    batch = {column: [] for column in columns}
                    batch_rows = 0

            if batch_rows:
                writer.write_table(pa.table(batch, schema=schema))
                wrote_rows = True
            if not wrote_rows:
                writer.write_table(
                    pa.Table.from_arrays(
                        [pa.array([], type=pa.string()) for _ in columns],
                        schema=schema,
                    )
                )
        finally:
            if workbook is not None:
                workbook.close()
            writer.close()
